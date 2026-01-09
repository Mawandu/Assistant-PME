# routers/data_source.py
import uuid
from fastapi import APIRouter, Depends, HTTPException, status, UploadFile, File
from sqlalchemy.orm import Session
from sqlalchemy import func, create_engine
from sqlalchemy.exc import OperationalError
from typing import List 
import json 
from encryption import encrypt_data, decrypt_data 
import models, schemas
from database import get_db
import time
from routers.auth import get_current_user, get_current_user_or_default

router = APIRouter(
    prefix="/api/v1/datasources",
    tags=['Data Sources']
)

def encrypt_connection_config(config: dict | None) -> dict | None:
    if config is None:
        return None
    encrypted_config = {}
    sensitive_keys = ["password", "secret", "token", "key", "pwd"] 
    for key, value in config.items():
        if any(sk in key.lower() for sk in sensitive_keys) and isinstance(value, str) and value:
            encrypted_config[key] = encrypt_data(value)
        else:
            encrypted_config[key] = value 
    return encrypted_config

def decrypt_connection_config(config: dict | None) -> dict | None:
    if config is None:
        return None
    decrypted_config = {}
    sensitive_keys = ["password", "secret", "token", "key", "pwd"]
    for key, value in config.items():
        if any(sk in key.lower() for sk in sensitive_keys) and isinstance(value, str):
            try:
                if value.startswith('gAAAAA'):
                    decrypted_config[key] = decrypt_data(value)
                else:
                     decrypted_config[key] = value 
            except ValueError:
                decrypted_config[key] = "DECRYPTION_FAILED"
        else:
            decrypted_config[key] = value
    return decrypted_config

@router.post("/{ds_id}/validate", response_model=schemas.DataSourceOut)
async def validate_data_source_connection(
    ds_id: uuid.UUID,
    db: Session = Depends(get_db),
    current_user: models.User = Depends(get_current_user_or_default)
):
    """
    Tente de se connecter à la source de données pour valider la configuration.
    Met à jour le statut de la source (ACTIVE ou ERROR).
    Nécessite une authentification valide (ou fallback dev).
    NOTE: Pour la production, ceci devrait être une tâche asynchrone (ex: Celery).
    """
    db_data_source = db.query(models.DataSource).filter(
        models.DataSource.id == ds_id,
        models.DataSource.tenant_id == current_user.tenant_id
    ).first()

    if db_data_source is None:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail=f"Data source with id {ds_id} not found."
        )

    connection_successful = False
    error_message = None

    try:
        config_to_test = decrypt_connection_config(db_data_source.connection_config)
        if config_to_test is None:
             raise ValueError("Connection config is missing.")
        if any(v == "DECRYPTION_FAILED" for v in config_to_test.values()):
             raise ValueError("Failed to decrypt sensitive connection details.")

    except ValueError as e:
         error_message = f"Configuration error: {e}"
         config_to_test = None 

    if config_to_test:
        try:
            if db_data_source.type == models.DataSourceType.SQL_CONNECTOR:
                db_user = config_to_test.get("username")
                db_password = config_to_test.get("password")
                db_host = config_to_test.get("host")
                db_port = config_to_test.get("port", 5432) 
                db_name = config_to_test.get("database")

                if not all([db_user, db_password, db_host, db_name]):
                    raise ValueError("Missing required connection details (username, password, host, database).")

                connection_url = f"postgresql://{db_user}:{db_password}@{db_host}:{db_port}/{db_name}"
                temp_engine = create_engine(connection_url, connect_args={'connect_timeout': 5}) 
                with temp_engine.connect() as connection:
                    connection_successful = True
                temp_engine.dispose() 

            elif db_data_source.type == models.DataSourceType.FILE_UPLOAD:
                connection_successful = True 
            else:
                error_message = f"Validation not implemented for type '{db_data_source.type.value}'."

        except OperationalError as e:
            error_message = f"Connection failed: {e.orig}" 
        except ValueError as e: 
             error_message = f"Configuration error: {e}"
        except Exception as e: 
            error_message = f"An unexpected error occurred: {e}"

    if connection_successful:
        db_data_source.status = models.DataSourceStatus.ACTIVE
        db_data_source.last_sync_status = "Validated"
        db_data_source.last_sync_error = None
    else:
        db_data_source.status = models.DataSourceStatus.ERROR
        db_data_source.last_sync_status = "Validation Failed"
        db_data_source.last_sync_error = str(error_message)[:255] 
    db_data_source.last_sync_at = func.now() 
    try:
        db.commit()
        db.refresh(db_data_source)
    except Exception as e:
        db.rollback()
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"Could not update data source status after validation: {e}"
        )

    if not connection_successful:
         raise HTTPException(
             status_code=status.HTTP_400_BAD_REQUEST,
             detail=f"Connection validation failed: {error_message}"
         )

    return db_data_source

@router.post("/", response_model=schemas.DataSourceOut, status_code=status.HTTP_201_CREATED)
async def create_data_source(
    ds_request: schemas.DataSourceCreate,
    db: Session = Depends(get_db),
    current_user: models.User = Depends(get_current_user_or_default) 
):
    """
    Crée une nouvelle source de données pour le tenant de l'utilisateur courant.
    Nécessite une authentification valide (ou fallback dev).
    """
    encrypted_config = encrypt_connection_config(ds_request.connection_config)
    # TODO: 

    new_ds = models.DataSource(
        name=ds_request.name,
        type=ds_request.type,
        sync_enabled=ds_request.sync_enabled,
        sync_frequency_minutes=ds_request.sync_frequency_minutes,
        connection_config=encrypted_config,
        tenant_id=current_user.tenant_id,
        status=models.DataSourceStatus.PENDING
    )
    try:
        db.add(new_ds)
        db.commit()
        db.refresh(new_ds)
    except Exception as e:
        db.rollback()
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"Could not create data source: {e}"
        )

    # TODO: 

    return new_ds

@router.put("/{ds_id}", response_model=schemas.DataSourceOut)
async def update_data_source(
    ds_id: uuid.UUID,
    ds_update_request: schemas.DataSourceUpdate,
    db: Session = Depends(get_db),
    current_user: models.User = Depends(get_current_user_or_default)
):
    """
    Met à jour une source de données. Crypte connection_config si fourni.
    """
    db_data_source = db.query(models.DataSource).filter(
        models.DataSource.id == ds_id,
        models.DataSource.tenant_id == current_user.tenant_id
    ).first()

    if db_data_source is None:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail=f"Data source {ds_id} not found")

    update_data = ds_update_request.dict(exclude_unset=True)

    # Si connection_config est mis à jour, le crypter
    if "connection_config" in update_data:
        update_data["connection_config"] = encrypt_connection_config(update_data["connection_config"])

    for key, value in update_data.items():
        setattr(db_data_source, key, value)

    # ... (try/except pour db.commit, db.refresh) ...
    try:
        db.commit()
        db.refresh(db_data_source)
    except Exception as e:
        db.rollback()
        # Logguez l'erreur e pour le débogage
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"Could not update data source: {e}"
        )
    return db_data_source

@router.get("/", response_model=List[schemas.DataSourceOut])
async def list_data_sources(db: Session = Depends(get_db), current_user: models.User = Depends(get_current_user_or_default)):
    """
    Liste toutes les sources de données pour le tenant de l'utilisateur courant.
    Nécessite une authentification valide (ou fallback dev).
    """
    data_sources = db.query(models.DataSource)\
        .filter(models.DataSource.tenant_id == current_user.tenant_id)\
        .order_by(models.DataSource.created_at.desc())\
        .all()
    return data_sources


@router.get("/{ds_id}", response_model=schemas.DataSourceOut)
async def get_data_source(
    ds_id: uuid.UUID,
    db: Session = Depends(get_db),
    current_user: models.User = Depends(get_current_user_or_default) 
):
    """
    Récupère les détails d'une source de données spécifique par son ID.
    Vérifie que la source appartient bien au tenant de l'utilisateur.
    Nécessite une authentification valide (ou fallback dev).
    """
    data_source = db.query(models.DataSource).filter(
        models.DataSource.id == ds_id,
        models.DataSource.tenant_id == current_user.tenant_id
    ).first()

    if data_source is None:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail=f"Data source with id {ds_id} not found."
        )
    return data_source

@router.put("/{ds_id}", response_model=schemas.DataSourceOut)
async def update_data_source(
    ds_id: uuid.UUID,
    ds_update_request: schemas.DataSourceUpdate,
    db: Session = Depends(get_db),
    current_user: models.User = Depends(get_current_user_or_default) 
):
    """
    Met à jour une source de données spécifique par son ID.
    Permet des mises à jour partielles (seuls les champs fournis sont modifiés).
    Vérifie que la source appartient bien au tenant de l'utilisateur.
    Nécessite une authentification valide (ou fallback dev).
    """
    # Récupérer la source de données existante
    db_data_source = db.query(models.DataSource).filter(
        models.DataSource.id == ds_id,
        models.DataSource.tenant_id == current_user.tenant_id
    ).first()

    if db_data_source is None:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail=f"Data source with id {ds_id} not found."
        )

    # Obtenir les données de la requête, en excluant les valeurs non définies (None)
    update_data = ds_update_request.dict(exclude_unset=True)

    # TODO: Ajouter la logique de cryptage si connection_config est mis à jour
    # TODO: Si le statut ou la config change, déclencher une re-validation ?

    # Mettre à jour les champs fournis dans l'objet SQLAlchemy
    for key, value in update_data.items():
        setattr(db_data_source, key, value)

    try:
        db.commit()
        db.refresh(db_data_source)
    except Exception as e:
        db.rollback()
        # Logguez l'erreur e pour le débogage
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"Could not update data source: {e}"
        )

    return db_data_source

@router.delete("/{ds_id}", status_code=status.HTTP_204_NO_CONTENT)
async def delete_data_source(
    ds_id: uuid.UUID,
    db: Session = Depends(get_db),
    current_user: models.User = Depends(get_current_user_or_default)
):
    """
    Supprime une source de données et le fichier associé si c'est un upload.
    """
    db_data_source = db.query(models.DataSource).filter(
        models.DataSource.id == ds_id,
        models.DataSource.tenant_id == current_user.tenant_id
    ).first()

    if db_data_source is None:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail=f"Data source with id {ds_id} not found."
        )

    # If it's a file upload, delete the file from disk
    if db_data_source.type == models.DataSourceType.FILE_UPLOAD:
        config = db_data_source.connection_config
        if config and "file_path" in config:
            import os
            file_path = config["file_path"]
            if os.path.exists(file_path):
                try:
                    os.remove(file_path)
                    print(f"DEBUG: Deleted file {file_path}")
                except Exception as e:
                    print(f"DEBUG: Error deleting file {file_path}: {e}")

    try:
        # Delete associated products (and their stock movements via cascade)
        # We use synchronize_session=False for bulk delete performance
        deleted_products = db.query(models.Product).filter(
            models.Product.data_source_id == ds_id
        ).delete(synchronize_session=False)
        
        print(f"DEBUG: Deleted {deleted_products} products associated with source {ds_id}")

        db.delete(db_data_source)
        db.commit()
    except Exception as e:
        db.rollback()
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"Could not delete data source: {e}"
        )
    
    return None

@router.post("/upload", status_code=status.HTTP_201_CREATED)
async def upload_file(
    file: UploadFile = File(...),
    db: Session = Depends(get_db),
    current_user: models.User = Depends(get_current_user_or_default)
):
    """
    Upload a file (Excel or CSV) to ingest data.
    Parses the file and creates/updates Products and StockMovements.
    (Optimized for Vercel: Uses openpyxl/csv instead of pandas)
    """
    print(f"DEBUG: Received file upload request: {file.filename}")
            
    # Imports
    import io
    import csv 
    try:
        from openpyxl import load_workbook
    except ImportError:
        # Should be in requirements
        print("WARNING: openpyxl not found")
        pass

    # 1. Read file content
    try:
        contents = await file.read()
        filename = file.filename.lower()
        print(f"DEBUG: Read {len(contents)} bytes from {filename}")
    except Exception as e:
        print(f"DEBUG: Error reading file: {e}")
        raise HTTPException(status_code=400, detail=f"Error reading file: {str(e)}")

    # 2. Save file to disk
    import os
    upload_dir = "uploads"
    if not os.path.exists(upload_dir):
        os.makedirs(upload_dir)
        
    file_path = os.path.join(upload_dir, f"{uuid.uuid4()}_{filename}")
    try:
        with open(file_path, "wb") as f:
            f.write(contents)
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error saving file: {str(e)}")

    # Create DataSource Record
    try:
        new_ds = models.DataSource(
            name=filename,
            type=models.DataSourceType.FILE_UPLOAD,
            connection_config={"file_path": file_path, "original_name": filename},
            tenant_id=current_user.tenant_id,
            status=models.DataSourceStatus.PENDING,
            last_sync_at=func.now()
        )
        db.add(new_ds)
        db.flush()
        
        # 3. Parse Data (No Pandas)
        rows = []
        
        column_mapping = {
            'product': 'name', 'product name': 'name', 'nom produit': 'name', 'nom': 'name', 'designation': 'name', 'libelle': 'name', 'product_name': 'name',
            'sku': 'sku', 'ref': 'sku', 'reference': 'sku', 'code': 'sku', 'product_id': 'sku',
            'category': 'category', 'catégorie': 'category', 'famille': 'category',
            'quantity': 'quantity', 'qty': 'quantity', 'quantité': 'quantity', 'stock': 'quantity', 'qte': 'quantity', 'stock reel': 'quantity', 'quantity_in_stock': 'quantity',
            'price': 'unit_price', 'prix': 'unit_price', 'unit price': 'unit_price', 'prix unitaire': 'unit_price', 'pamp': 'cost_price', 'coût': 'cost_price', 'cost': 'cost_price',
            'supplier': 'supplier_name', 'fournisseur': 'supplier_name'
        }

        def normalize_header(h):
            return str(h).lower().strip()

        def map_header(h):
            norm = normalize_header(h)
            return column_mapping.get(norm, norm)

        if filename.endswith('.csv'):
            # Handle CSV
            try:
                content_str = contents.decode('utf-8-sig') # Handle BOM
            except UnicodeDecodeError:
                content_str = contents.decode('latin-1') # Fallback

            f_io = io.StringIO(content_str)
            reader = csv.DictReader(f_io)
            
            # Remap fieldnames
            if reader.fieldnames:
                reader.fieldnames = [map_header(h) for h in reader.fieldnames]
                rows = list(reader)
                
        elif filename.endswith(('.xls', '.xlsx')):
            # Handle Excel
            wb = load_workbook(io.BytesIO(contents), data_only=True)
            ws = wb.active
            
            # Extract headers from first row
            header_row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True), None)
            if header_row:
                headers = [map_header(h) if h is not None else f"col_{i}" for i, h in enumerate(header_row)]
                
                for row_values in ws.iter_rows(min_row=2, values_only=True):
                    # Create dict
                    row_dict = {headers[i]: val for i, val in enumerate(row_values) if i < len(headers)}
                    rows.append(row_dict)
        else:
            raise HTTPException(status_code=400, detail="Invalid file format.")

        # Process logic
        processed_count = 0
        errors = []

        for index, row in enumerate(rows):
            try:
                # Helper for safe access
                def get_val(key):
                    v = row.get(key)
                    if v is None or v == "": return None
                    return v

                product_name = get_val('name')
                sku = get_val('sku')
                
                if not sku and not product_name: continue
                if not sku: continue

                # Ensure Category
                category_name = get_val('category') or 'General'
                
                category = db.query(models.Category).filter(
                     models.Category.name == str(category_name),
                     models.Category.tenant_id == current_user.tenant_id
                ).first()
                if not category:
                     category = models.Category(name=str(category_name), tenant_id=current_user.tenant_id)
                     db.add(category)
                     db.flush()

                # Supplier
                supplier_id = None
                supplier_name = get_val('supplier_name')
                if supplier_name:
                     supplier = db.query(models.Supplier).filter(models.Supplier.name == str(supplier_name), models.Supplier.tenant_id == current_user.tenant_id).first()
                     if not supplier:
                         supplier = models.Supplier(name=str(supplier_name), tenant_id=current_user.tenant_id)
                         db.add(supplier)
                         db.flush()
                     supplier_id = supplier.id

                # Product
                product = db.query(models.Product).filter(
                    models.Product.sku == str(sku),
                    models.Product.tenant_id == current_user.tenant_id
                ).first()

                price = get_val('unit_price')
                cost = get_val('cost_price')
                
                # Conversion helpers
                def to_float(x):
                    try: return float(x)
                    except: return 0.0
                
                if not product:
                    product = models.Product(
                        sku=str(sku),
                        name=str(product_name) if product_name else f"Product {sku}",
                        category_id=category.id,
                        unit_price=to_float(price),
                        cost_price=to_float(cost),
                        supplier_id=supplier_id,
                        tenant_id=current_user.tenant_id,
                        data_source_id=new_ds.id
                    )
                    db.add(product)
                    db.flush()
                
                # Stock Movement
                qty = get_val('quantity')
                if qty is not None:
                     try:
                         q = int(qty)
                         if q > 0:
                             mv = models.StockMovement(
                                 product_id=product.id,
                                 movement_type="INBOUND",
                                 quantity=q,
                                 notes=f"Import via {filename}",
                                 user_id=current_user.id,
                                 tenant_id=current_user.tenant_id
                             )
                             db.add(mv)
                     except: pass
                
                processed_count += 1
            except Exception as inner_e:
                errors.append(f"Row {index}: {str(inner_e)}")

        new_ds.status = models.DataSourceStatus.ACTIVE
        new_ds.last_sync_status = "COMPLETED"
        db.commit()
        
    except Exception as e:
        db.rollback()
        print(f"Upload failed: {e}")
        raise HTTPException(status_code=500, detail=f"Upload failed: {str(e)}")

    message = f"File '{filename}' processed. "
    if processed_count > 0:
        message += f"{processed_count} items."
    
    return {
        "message": message,
        "data_source_id": str(new_ds.id),
        "errors": errors[:10]
    }